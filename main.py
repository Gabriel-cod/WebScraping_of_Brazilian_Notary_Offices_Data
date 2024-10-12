from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.common.exceptions import *
from selenium.webdriver.support import expected_conditions
from time import sleep
from selenium.webdriver.support.select import Select
import openpyxl

class ScrapyDados():
    def __init__(self):
        chrome_options = Options()
        arguments = ['--lang=pt-BR', '--start-maximized', '--incognito', '--disable-notifications']
        for argument in arguments:
            chrome_options.add_argument(argument)

        chrome_options.add_experimental_option('prefs', {
            'download.prompt_for_download': False,
            'profile.default_content_setting_values.notifications': 2,
            'profile.default_content_setting_values.automatic_downloads': 1,
        })
        self.driver = webdriver.Chrome(options=chrome_options)

        self.wait = WebDriverWait(
            driver=self.driver,
            timeout=10,
            poll_frequency=1,
            ignored_exceptions=[
                ElementNotVisibleException,
                NoSuchElementException,
                ElementNotSelectableException,
                ElementNotInteractableException
            ]
        )

    def acessar_estado(self, estado, url):
        self.driver.get(url=url)
        try:
            avancado_button = self.wait.until(expected_conditions.element_to_be_clickable((By.ID, "details-button")))
            avancado_button.click(); sleep(1.5)
            self.driver.execute_script('window.scrollTo(0, document.body.scrollHeight);')
            continuar_button = self.wait.until(expected_conditions.element_to_be_clickable((By.ID, "proceed-link")))
            continuar_button.click()
        except:
            pass

        click_button = self.wait.until(expected_conditions.visibility_of_any_elements_located((By.LINK_TEXT, "Clique aqui")))
        click_button[0].click()

        sleep(2)
        self.driver.execute_script('window.scrollTo(0, document.body.scrollHeight);'); sleep(2)
        while True:
            if len(estado) > 2:
                try:
                    estado_element = self.driver.find_element(By.XPATH, f"//area[@title='{estado.capitalize()}']")
                    self.driver.execute_script('arguments[0].click()', estado_element)
                    break
                except:
                    ScrapyDados.formatar_titulo('ATENÇÃO: Nome ou sigla do estado não encontrado. Verifique se foi inserido o nome correto e tente novamente.')
                    estado = str(input('Informe o ESTADO onde a cidade se situa [sigla ou nome completo e acentuado] >>> ')).strip()
            else:
                print(estado, len(estado))
                try:
                    estado_element = self.driver.find_element(By.XPATH, f'//area[@onclick="pesquisaServentiasExtra(\'{str(estado).upper()}\')"]')
                    self.driver.execute_script('arguments[0].click()', estado_element)
                    break
                except:
                    ScrapyDados.formatar_titulo('ATENÇÃO: Nome ou sigla do estado não encontrado. Verifique se foi inserido o nome correto e tente novamente.')
                    estado = str(input('Informe o ESTADO onde a cidade se situa [sigla ou nome completo e acentuado] >>> '))

    def acessar_cidade(self, cidade):
        select_campo = self.wait.until(expected_conditions.element_to_be_clickable((By.XPATH, "//select[@id='cidade_serventia']")))
        select_cidade = Select(select_campo)
        while True:
            try:
                select_cidade.select_by_visible_text(cidade)
                break
            except:
                ScrapyDados.formatar_titulo(f'ATENÇÃO: Cidade [{cidade}] não encontrada. Verifique se a cidade inserida está escrita corretamente e tente novamente.')
                cidade = str(input('Informe o nome da cidade >>> ')).strip().upper()

        pesquisar_button = self.driver.find_elements(By.XPATH, "//button[@type='submit']")[-1]
        pesquisar_button.click()
        return cidade

    def extrair_dados(self, planilha, cidade):
        quantidade_field = self.wait.until(expected_conditions.element_to_be_clickable((By.NAME, "display_length")))
        select_quantidade = Select(quantidade_field)
        select_quantidade.select_by_index(3); sleep(2)
        self.driver.execute_script('window.scrollTo(0, document.body.scrollHeight);'); sleep(2)
        self.driver.execute_script('window.scrollTo(0, document.body.scrollTop);'); sleep(1)
        registros_cns = self.wait.until(expected_conditions.visibility_of_all_elements_located((By.XPATH, "//table[@id='display']//tbody[@aria-live]/tr//td//strong/font[@color='red']")))
        registros_infos = self.wait.until(expected_conditions.visibility_of_all_elements_located((By.XPATH, "//table[@id='display']//tbody[@aria-live]/tr//td[2]//table//tbody//tr//td[@align='left']")))
        info = 0
        lista_linhas = []
        while info <= len(registros_infos):
            try:
                nome = registros_infos[info].text
            except IndexError:
                break
            responsavel = registros_infos[info+1].text
            atributo = registros_infos[info+2].text
            telefone_e_email = str(registros_infos[info+4].text)
            telefone = telefone_e_email.split('E-mail :')[0].strip()
            email = telefone_e_email.split('E-mail :')[1].strip()
            cns = registros_cns[(info+1)//5].text
            lista_infos = [cidade, cns, str(nome).strip(), str(responsavel).capitalize().strip(), str(telefone).strip(), str(email).strip(), str(atributo).capitalize().strip()]
            lista_linhas.append(lista_infos)
            info += 5
        ScrapyDados.armazenar_dados(linhas=lista_linhas, wb_name=planilha, cidade=cidade)
        sleep(2)

    @staticmethod
    def armazenar_dados(linhas, wb_name, cidade):
        try:
            book = openpyxl.load_workbook(f'{wb_name}.xlsx')
        except:
            book = openpyxl.Workbook()
        try:
            city_page = book[f'{cidade}']
        except:
            book.create_sheet(f'{cidade}')
            city_page = book[f'{cidade}']
            city_page.append(['Cidade', 'CNS', 'Nome', 'Responsável', 'Telefone', 'E-mail', 'Atribuição'])
        try:
            del book['Sheet']
        except:
            pass
        for linha in linhas:
            city_page.append(linha)
        book.save(f'{wb_name}.xlsx')

    @staticmethod
    def formatar_titulo(text):
        print('-'*(len(text)+4))
        print(f'  {text}')
        print('-'*(len(text)+4))

ScrapyDados.formatar_titulo('Bem-vindo à Busca por Cartórios')
site = "https://www.cnj.jus.br/corregedoria/justica_aberta/?"
while True:
    estado = str(input('Informe o ESTADO onde a cidade se situa [sigla ou nome completo e acentuado] >>> ')).strip()
    cidades = str(input('''Informe os nomes das CIDADES do mesmo estado, separados por ", "
Ex.: Belo Horizonte, Curvelo, Varginha, Diamantina  >>> ''')).strip().upper().split(', ')
    nome_planilha = str(input('''Informe o nome da planilha onde serão armazenados esses dados.
Obs.: Caso queira armazenar os dados em uma planilha já existente na pasta, apenas informe o nome exato dela.
Informe o nome sem a extensão ".XLSX" e sem espaços no nome. Ex.: Cartorios_Minas_Gerais

>>> '''))
    ScrapyDados.formatar_titulo('Aguarde a extração.')

    scrape_cartorios_bot = ScrapyDados()
    for cidade in cidades:
        scrape_cartorios_bot.acessar_estado(estado=estado, url=site)
        cidade = scrape_cartorios_bot.acessar_cidade(cidade=cidade)
        scrape_cartorios_bot.extrair_dados(cidade=cidade, planilha=nome_planilha)
        # scrape_cartorios_bot.driver.close()
    ScrapyDados.formatar_titulo('Finalizado.')
    while True:
        continuar = str(input('Deseja fazer outra extração? [s/n] >>> ')).strip().lower()[0]
        if continuar in ['n', 's']:
            break
        else:
            ScrapyDados.formatar_titulo('ATENÇÃO: digite somente [s] para sim ou [n] não.')

    if continuar == 'n':
        break
